import os
import sys
import time
import glob
import shutil
from uuid import uuid4
from datetime import datetime, timedelta, timezone
from urllib.parse import urlencode

# --- Gestión de Variables de Entorno ---
from dotenv import load_dotenv

# --- Automatización web con Playwright (Microsoft Edge / Chromium) ---
from playwright.sync_api import (
    BrowserContext,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)

# --- Manejo de Excel (Openpyxl) ---
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- Comunicación con Power Automate (Webhooks y Red) ---
import base64
import csv
import io
import json
import re
import unicodedata
import zipfile
import pandas as pd
import requests
import psycopg
from psycopg import sql


# =========================
#  Config consolidado
# =========================
BBDD_HEADERS = [
    "Resource", "Nombre", "Nombre Contacto", "Date", "Activity Type",
    "Lugar de la Reparación", "Reparación en Nodo", "Reparación en DC", "Activity Status",
    "Identificador (número de reclamo)", "ID de Ticket", "Work Order", "Start", "End", "Start - End",
    "Traveling Time", "Duration", "Name", "idRuta", "Address", "City", "State", "Telefono", "Correo",
    "Account Number", "Work Zone", "Activity ID", "Activity Time of Booking",
    "Motivo de suspensión de actividad", "Motivo de cancelación de activida", "FIIL", "Tipo de Fiil",
    "Recurso Logico", "Metraje", "Zona", "ID_Servicio1", "ID_Servicio2", "ID_Servicio3", "ID_Servicio4",
    "ID_Servicio5", "ID_Servicio6", "ID_Servicio7", "ID_Servicio8", "ID_Servicio9", "ID_Servicio10",
    "ID Servicio", "CS Enlace", "CS Equipo", "CS Telefonía", "Número de OT", "Estado de OT",
    "Identificador del Reclamo Masivo", "Área resolutora", "Descripción del servicio", "Folio Externo",
    "Tipo de cliente", "Pre Clasificación Detalle", "Prioridad", "Time Slot", "Folio Subtel",
    "Tipo de reclamo masivo", "Tipo de servicio", "Cable", "Filamento", "Nivel1", "Nivel2", "Nivel3",
    "Razón de completación", "Intervención Planta Externa", "Ticket de Trazabilidad",
    "Causa de la falla", "Area elemento", "Solucion", "Comuna", "Subgrupo TOA", "Zona Subgrupo TOA"
]

ACTIVITY_TYPE_ALLOWED_VALUES = [
    "Certificación",
    "Certificación de enlace-Fibra",
    "Construcción + Habilitación",
    "Construcción de enlace",
    "Habilitación",
    "Habilitación + certificación",
    "Habilitación de servicio",
]


# =========================
#  Fecha/hora
# =========================
def get_tz_scl():
    try:
        from zoneinfo import ZoneInfo
        try:
            return ZoneInfo("America/Santiago")
        except Exception:
            import tzdata  # noqa
            return ZoneInfo("America/Santiago")
    except Exception:
        pass
    try:
        return datetime.now().astimezone().tzinfo
    except Exception:
        pass
    return timezone(timedelta(hours=-3))


def now_cl():
    return datetime.now(get_tz_scl())


def human_dt(dt: datetime):
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# =========================
#  FS helpers
# =========================
def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def atomic_move_replace(src: str, dest: str):
    dest_dir = os.path.dirname(dest)
    ensure_dir(dest_dir)
    tmp = os.path.join(dest_dir, f".tmp_{uuid4().hex}_{os.path.basename(dest)}")
    shutil.move(src, tmp)
    os.replace(tmp, dest)


def _remove_files(paths: list[str], log_prefix: str, retries: int = 5) -> list[str]:
    """Elimina archivos concretos, con reintentos y registro por archivo."""
    errors = []

    for path in dict.fromkeys(os.path.abspath(path) for path in paths):
        if not os.path.isfile(path):
            continue

        last_error = None
        for attempt in range(1, retries + 1):
            try:
                os.remove(path)
                print(f"{log_prefix} Eliminado: {path}")
                last_error = None
                break
            except OSError as exc:
                last_error = exc
                if attempt < retries:
                    time.sleep(0.5 * attempt)

        if last_error is not None:
            error = f"No se pudo eliminar '{path}': {last_error!r}"
            errors.append(error)
            print(f"{log_prefix} ERROR: {error}", file=sys.stderr)

    return errors


def remove_old_download_files(
    download_dir: str,
    protected_paths: list[str] | None = None,
) -> None:
    """Limpia los XLSX y temporales dejados por ejecuciones anteriores."""
    protected = {
        os.path.normcase(os.path.abspath(path))
        for path in (protected_paths or [])
        if path
    }
    old_files = []
    for entry in os.scandir(download_dir):
        if not entry.is_file(follow_symlinks=False):
            continue
        normalized_path = os.path.normcase(os.path.abspath(entry.path))
        if normalized_path in protected:
            continue
        if entry.name.lower().endswith(".xlsx") or entry.name.startswith(".tmp_"):
            old_files.append(entry.path)

    if not old_files:
        print(f"[LIMPIEZA INICIAL] No hay archivos antiguos en: {download_dir}")
        return

    errors = _remove_files(old_files, "[LIMPIEZA INICIAL]")
    if errors:
        raise RuntimeError(
            "La limpieza inicial no pudo completarse:\n" + "\n".join(errors)
        )


def _available_download_path(
    download_dir: str,
    server_file_name: str,
    fallback_name: str,
) -> str:
    """Genera un destino que nunca sobrescribe otra descarga de la ejecución."""
    destination = os.path.join(download_dir, server_file_name)
    if not os.path.exists(destination):
        return destination

    fallback_destination = os.path.join(download_dir, fallback_name)
    if not os.path.exists(fallback_destination):
        return fallback_destination

    stem, extension = os.path.splitext(fallback_name)
    suffix = 2
    while True:
        candidate = os.path.join(download_dir, f"{stem}_{suffix}{extension}")
        if not os.path.exists(candidate):
            return candidate
        suffix += 1


# =========================
#  Excel helpers (Monitoreo)
# =========================
def append_to_table_atomic(
    xlsx_path: str,
    sheet_name: str,
    table_name: str,
    row_values: list,
    retries: int = 3
):
    for _ in range(retries):
        try:
            wb = load_workbook(xlsx_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

            tbl = None
            for t in ws.tables.values():
                if t.displayName == table_name:
                    tbl = t
                    break
            if tbl is None:
                wb.close()
                raise ValueError(f"No se encontró la tabla '{table_name}' en {xlsx_path}")

            min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
            next_row = max_row + 1

            for idx, val in enumerate(row_values):
                col_letter = get_column_letter(min_col + idx)
                ws[f"{col_letter}{next_row}"] = val

            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{next_row}"
            tbl.ref = new_ref

            tmp = f"{xlsx_path}.tmp"
            wb.save(tmp)
            wb.close()
            os.replace(tmp, xlsx_path)
            return
        except PermissionError:
            time.sleep(2)
        except Exception:
            raise


def update_last_row_fields_atomic(
    xlsx_path: str,
    sheet_name: str,
    table_name: str,
    field_values: dict
):
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    tbl = None
    for t in ws.tables.values():
        if t.displayName == table_name:
            tbl = t
            break
    if tbl is None:
        wb.close()
        raise ValueError(f"No se encontró la tabla '{table_name}' en {xlsx_path}")

    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)

    headers = []
    for j in range(min_col, max_col + 1):
        headers.append(ws.cell(row=min_row, column=j).value or "")

    col_index = {}
    for field in field_values.keys():
        if field in headers:
            col_index[field] = headers.index(field) + min_col
        else:
            max_col += 1
            new_col_letter = get_column_letter(max_col)
            ws[f"{new_col_letter}{min_row}"] = field
            headers.append(field)
            col_index[field] = max_col

    target_row = max_row
    for field, value in field_values.items():
        ws.cell(row=target_row, column=col_index[field], value=value)

    end_col_letter = get_column_letter(max_col)
    tbl.ref = f"{get_column_letter(min_col)}{min_row}:{end_col_letter}{max_row}"

    tmp = f"{xlsx_path}.tmp"
    wb.save(tmp)
    wb.close()
    os.replace(tmp, xlsx_path)


# =========================
#  Lectura de InputsTOA.xlsx
# =========================
def load_inputs_toa(ruta_inputs: str) -> list[dict]:
    """
    Lee el archivo InputsTOA.xlsx y retorna una lista de dicts con las
    columnas relevantes, filtrando solo las filas con ESTADO_EN_SCRIPT == "ACTIVO".

    Columnas consideradas:
        "Nombre Archivo Completo", "Subgrupo TOA", "Zona Subgrupo TOA",
        "URL", "RUTA_Carpeta_Destino", "ESTADO_EN_SCRIPT",
        "providerId", "downloadId", "recursively"
    """
    COLS_NEEDED = {
        "Nombre Archivo Completo",
        "Subgrupo TOA",
        "Zona Subgrupo TOA",
        "URL",
        "RUTA_Carpeta_Destino",
        "ESTADO_EN_SCRIPT",
        "providerId",
        "downloadId",
        "recursively",
    }

    if not os.path.isfile(ruta_inputs):
        raise FileNotFoundError(f"No se encontró el archivo InputsTOA.xlsx en: {ruta_inputs}")

    wb = load_workbook(ruta_inputs, data_only=True, read_only=True)
    ws = wb.active

    # Leer encabezados de la primera fila
    raw_headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_map = {h: i for i, h in enumerate(raw_headers) if h in COLS_NEEDED}

    missing_cols = COLS_NEEDED - set(col_map.keys())
    if missing_cols:
        wb.close()
        raise ValueError(
            f"Faltan columnas en InputsTOA.xlsx: {', '.join(sorted(missing_cols))}"
        )

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue

        estado = str(row[col_map["ESTADO_EN_SCRIPT"]] or "").strip().upper()
        if estado != "ACTIVO":
            continue

        record = {col: (str(row[idx]).strip() if row[idx] is not None else "") for col, idx in col_map.items()}
        records.append(record)

    wb.close()
    print(f"[INPUTS] {len(records)} registros ACTIVOS cargados desde InputsTOA.xlsx")
    return records


def build_toa_url(record: dict, date_str: str) -> str:
    """
    Construye la URL de descarga según la lógica de concatenación definida:
    part1_URL + providerId + part1date_URL + date + part2_URL + downloadId
    + part2date_URL + date + part3_URL + recursively
    """
    part1_url      = "https://gtd-zcn.etadirect.com/?m=gridexport&a=download&itype=manage&providerId="
    part1date_url  = "&date="
    part2_url      = "&panel=top&view=time&downloadId="
    part2date_url  = "&dates="
    part3_url      = "&recursively="

    return (
        f"{part1_url}{record['providerId']}"
        f"{part1date_url}{date_str}"
        f"{part2_url}{record['downloadId']}"
        f"{part2date_url}{date_str}"
        f"{part3_url}{record['recursively']}"
    )


# =========================
#  Consolidado helpers
# =========================
def _clear_sheet_and_ensure_table(wb, sheet_name: str, table_name: str, headers: list):
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)

    existing = None
    for t in ws.tables.values():
        if t.displayName == table_name:
            existing = t
            break

    end_col_letter = get_column_letter(len(headers))
    if existing is None:
        new_ref = f"A1:{end_col_letter}1"
        tbl = Table(displayName=table_name, ref=new_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)
    else:
        existing.ref = f"A1:{end_col_letter}1"

    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)

    return ws


def _append_bulk_rows_and_resize_table(
    wb,
    sheet_name: str,
    table_name: str,
    rows: list,
    headers_count: int
):
    ws = wb[sheet_name]
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    end_row = max(1 + len(rows), 1)
    end_col_letter = get_column_letter(headers_count)
    new_ref = f"A1:{end_col_letter}{end_row}"

    tbl = None
    for t in ws.tables.values():
        if t.displayName == table_name:
            tbl = t
            break
    if tbl is None:
        raise ValueError(f"No se encontró la tabla '{table_name}' para redimensionar.")
    tbl.ref = new_ref


def _get_unique_dates_to_delete(rows: list, headers: list):
    """
    Identifica todas las fechas únicas de actividad presentes en las nuevas filas.
    """
    date_header = "Date"
    try:
        date_col_idx = headers.index(date_header)
    except ValueError:
        print(f"[ERROR] La cabecera '{date_header}' no se encontró en {headers}", file=sys.stderr)
        return set()

    unique_dates = set()
    for row in rows:
        if date_col_idx < len(row):
            date_value = row[date_col_idx]
            if isinstance(date_value, datetime):
                unique_dates.add(date_value.date())
            elif isinstance(date_value, str) and date_value.strip():
                try:
                    dt = datetime.strptime(date_value.split(' ')[0], '%Y-%m-%d')
                    unique_dates.add(dt.date())
                except Exception:
                    pass
    return unique_dates


def _delete_rows_by_date_bulk(xlsx_path: str, sheet_name: str, table_name: str, dates_to_delete: set, headers: list):
    """
    Función atómica para eliminar todas las filas que coincidan con CUALQUIERA
    de las fechas de actividad en el set dates_to_delete.
    """
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    tbl = ws.tables.get(table_name)
    if tbl is None:
        wb.close()
        return 0

    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)

    current_headers = [ws.cell(row=min_row, column=j).value or "" for j in range(min_col, max_col + 1)]
    try:
        date_col_idx = current_headers.index("Date") + min_col
    except ValueError:
        wb.close()
        print(f"[ERROR] Columna 'Date' no encontrada en BBDD_TOA.", file=sys.stderr)
        return 0

    rows_to_delete = []
    deleted_count = 0

    for i in range(min_row + 1, max_row + 1):
        date_cell_value = ws.cell(row=i, column=date_col_idx).value
        current_date_obj = None
        if isinstance(date_cell_value, datetime):
            current_date_obj = date_cell_value.date()
        if current_date_obj in dates_to_delete:
            rows_to_delete.append(i)

    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num, 1)
        deleted_count += 1

    new_max_row = max_row - deleted_count
    end_col_letter = get_column_letter(max_col)

    if new_max_row <= min_row:
        new_ref = f"{get_column_letter(min_col)}{min_row}:{end_col_letter}{min_row}"
    else:
        new_ref = f"{get_column_letter(min_col)}{min_row}:{end_col_letter}{new_max_row}"

    tbl.ref = new_ref

    tmp = f"{xlsx_path}.tmp_del"
    wb.save(tmp)
    wb.close()
    os.replace(tmp, xlsx_path)

    return deleted_count


def _iter_source_rows(xlsx_path: str, subgroup_value: str, zona_subgrupo_value: str, headers: list):
    """
    Lee las filas de un archivo fuente y las mapea a la estructura de BBDD_HEADERS.
    Rellena "Subgrupo TOA" y "Zona Subgrupo TOA" con los valores del InputsTOA.xlsx.
    Retorna lista vacía si el archivo solo tiene encabezados (sin datos de valor).
    """
    rows = []
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active

    src_headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    name_to_idx = {}
    for i, h_name in enumerate(src_headers):
        if h_name not in name_to_idx:
            name_to_idx[h_name] = i

    data_started = False
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(cell is None for cell in r):
            continue
        data_started = True

        row_vals = []
        for h in headers:
            if h == "Subgrupo TOA":
                row_vals.append(subgroup_value)
            elif h == "Zona Subgrupo TOA":
                row_vals.append(zona_subgrupo_value)
            else:
                idx = name_to_idx.get(h)
                if idx is not None and idx < len(r):
                    row_vals.append(r[idx])
                else:
                    row_vals.append(None)

        rows.append(row_vals)

    wb.close()
    return rows if data_started else []


def consolidate_bbdd_toa(ruta_bbdd_toa: str, inputs_records: list, headers: list):
    
    print("[CONSOLIDADO] Iniciando…")
    if not ruta_bbdd_toa:
        raise RuntimeError("Falta variable .env RUTA_BBDD_TOA")

    all_new_rows = []
    files_count = 0

    folder_meta: dict[str, tuple[str, str]] = {}
    for rec in inputs_records:
        carpeta = rec["RUTA_Carpeta_Destino"]
        if carpeta:
            folder_meta[carpeta] = (rec["Subgrupo TOA"], rec["Zona Subgrupo TOA"])

    rows_added_counts: dict[str, int] = {}

    for folder, (subgroup, zona_subgrupo) in folder_meta.items():
        if not folder or not os.path.isdir(folder):
            print(f"[CONSOLIDADO] Carpeta no encontrada o inválida: {folder}")
            continue

        files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
        print(f"[CONSOLIDADO] {subgroup} ({zona_subgrupo}) → {len(files)} archivos en {folder}")

        for f in files:
            rows = _iter_source_rows(f, subgroup, zona_subgrupo, headers)
            if rows:
                all_new_rows.extend(rows)
                files_count += 1
                rows_added_counts[subgroup] = rows_added_counts.get(subgroup, 0) + len(rows)
            # Archivos con solo encabezados (rows vacío) se cuentan pero no se agregan (punto 9)

    if not all_new_rows:
        print("[CONSOLIDADO] No hay filas nuevas para consolidar. Listo.")
        return files_count, 0

    dates_to_replace = _get_unique_dates_to_delete(all_new_rows, headers)

    if not dates_to_replace:
        print("[CONSOLIDADO] No se encontraron fechas válidas en los nuevos archivos. Agregando sin limpieza...")
        rows_deleted = 0
    else:
        print(f"[CONSOLIDADO] Fechas únicas para reemplazo: {[d.strftime('%Y-%m-%d') for d in dates_to_replace]}")
        try:
            rows_deleted = _delete_rows_by_date_bulk(ruta_bbdd_toa, "BBDD_TOA", "BBDD_TOA", dates_to_replace, headers)
            print(f"[CONSOLIDADO] Filas antiguas eliminadas: {rows_deleted}")
        except Exception as e:
            print(
                f"[ERROR CONSOLIDADO] Falló la eliminación por fecha: {repr(e)}. "
                "Continuaremos, pero puede haber duplicados.",
                file=sys.stderr,
            )
            rows_deleted = 0

    sheet_name = "BBDD_TOA"
    table_name = "BBDD_TOA"

    wb = load_workbook(ruta_bbdd_toa) if os.path.exists(ruta_bbdd_toa) else Workbook()
    ws = _clear_sheet_and_ensure_table(wb, sheet_name, table_name, headers)

    print(f"[CONSOLIDADO] Escribiendo {len(all_new_rows)} filas nuevas…")
    _append_bulk_rows_and_resize_table(wb, sheet_name, table_name, all_new_rows, len(headers))
    rows_count = ws.max_row - 1

    tmp = f"{ruta_bbdd_toa}.tmp"
    wb.save(tmp)
    wb.close()
    os.replace(tmp, ruta_bbdd_toa)

    print("\n[CONSOLIDADO] Listo.")
    print(f"Modificaciones en BBDD_TOA (Total de Filas: {rows_count}):")
    print(f"  ❌ Filas Antiguas Eliminadas (Total): {rows_deleted}")
    print(f"  ✅ Filas Nuevas Agregadas:")
    for sg, cnt in rows_added_counts.items():
        print(f"    - {sg}: {cnt}")

    return files_count, rows_count


# =========================
#  Navegación / Playwright
# =========================
def xpath_locator(page: Page, xpath: str):
    """Crea un locator Playwright a partir de un XPath absoluto o relativo."""
    return page.locator(f"xpath={xpath}")


def find_signin_button_enabled(
    page: Page,
    primary_xpath: str,
    alternative_xpath: str,
    timeout_ms: int = 12_000,
):
    """Busca el primer botón de ingreso visible y habilitado."""
    candidates = []
    if primary_xpath:
        candidates.append(xpath_locator(page, primary_xpath.rstrip("/")))
    if alternative_xpath:
        candidates.append(xpath_locator(page, alternative_xpath.rstrip("/")))

    candidates.extend(
        [
            page.locator("xpath=//form//button[@type='submit' and not(@disabled)]"),
            page.locator(
                "xpath=//form//button[not(@disabled) "
                "and not(contains(@class,'oj-disabled'))]"
            ),
            page.locator(
                "xpath=//button[.//span[normalize-space()='Sign in'] "
                "or .//span[normalize-space()='Iniciar sesión'] "
                "or normalize-space()='Sign in' "
                "or normalize-space()='Iniciar sesión']"
            ),
        ]
    )

    deadline = time.monotonic() + (timeout_ms / 1000)
    last_error = None
    while time.monotonic() < deadline:
        for candidate in candidates:
            locator = candidate.first
            try:
                if locator.count() == 0:
                    continue
                if locator.is_visible() and locator.is_enabled():
                    class_name = locator.get_attribute("class") or ""
                    disabled = locator.get_attribute("disabled")
                    if disabled is None and "oj-disabled" not in class_name:
                        return locator
            except Exception as exc:
                last_error = exc
        page.wait_for_timeout(200)

    raise PlaywrightTimeoutError(
        f"Botón Sign in no habilitado dentro de {timeout_ms} ms: {last_error!r}"
    )


def submit_login(
    page: Page,
    password_locator,
    primary_xpath: str,
    alternative_xpath: str,
) -> None:
    """Envía el formulario usando click, Enter y submit() como alternativas."""
    try:
        button = find_signin_button_enabled(
            page,
            primary_xpath,
            alternative_xpath,
            timeout_ms=10_000,
        )
        button.click(force=True)
        return
    except Exception:
        pass

    try:
        password_locator.press("Enter")
        return
    except Exception:
        pass

    form = page.locator("form").first
    form.wait_for(state="attached", timeout=5_000)
    form.evaluate("form => form.submit()")


def wait_for_session_ready(
    page: Page,
    ready_xpath: str,
    timeout_ms: int = 120_000,
) -> None:
    """Espera cualquiera de las señales que indican que TOA terminó el login."""
    configured_ready = xpath_locator(page, ready_xpath)
    avatar_ready = page.locator(
        "xpath=//header//*[name()='visuals:technician-avatar']"
    )
    configured_ready.or_(avatar_ready).first.wait_for(
        state="attached",
        timeout=timeout_ms,
    )


def login_toa(
    page: Page,
    username: str,
    password: str,
    user_xpath: str,
    password_xpath: str,
    button_xpath: str,
    button_xpath_alt: str,
    ready_xpath: str,
) -> None:
    """Inicia sesión en TOA con reintentos y manejo de sesión anterior."""
    page.goto(
        "https://gtd-zcn.etadirect.com/",
        wait_until="domcontentloaded",
        timeout=120_000,
    )
    print("[LOGIN] Cargando login…")

    user_locator = xpath_locator(page, user_xpath).first
    password_locator = xpath_locator(page, password_xpath).first
    user_locator.wait_for(state="attached", timeout=60_000)
    password_locator.wait_for(state="attached", timeout=60_000)

    user_locator.fill(username)
    password_locator.fill(password)
    submit_login(page, password_locator, button_xpath, button_xpath_alt)

    # Oracle Field Service puede solicitar cerrar una sesión anterior.
    try:
        checkbox = page.locator("#delsession").first
        checkbox.wait_for(state="attached", timeout=20_000)
        try:
            if checkbox.get_attribute("type") == "checkbox":
                checkbox.check(force=True)
            else:
                checkbox.click(force=True)
        except Exception:
            checkbox.click(force=True)

        page.wait_for_timeout(400)
        submit_login(page, password_locator, button_xpath, button_xpath_alt)
        page.wait_for_timeout(600)

        password_locator = xpath_locator(page, password_xpath).first
        password_locator.wait_for(state="attached", timeout=15_000)
        password_locator.fill(password)
        submit_login(page, password_locator, button_xpath, button_xpath_alt)
        print("[LOGIN] Reenvío final con cierre de sesión anterior.")
    except PlaywrightTimeoutError:
        print("[LOGIN] No apareció el control de sesión anterior; flujo normal.")

    wait_for_session_ready(page, ready_xpath)

    # La interfaz puede estar adjunta al DOM antes de que la cookie de sesión
    # y las llamadas iniciales hayan terminado de estabilizarse.
    try:
        page.wait_for_load_state("domcontentloaded", timeout=30_000)
    except Exception:
        pass
    page.wait_for_timeout(2_000)

    session_cookies = page.context.cookies(
        ["https://gtd-zcn.etadirect.com/"]
    )
    if not session_cookies:
        raise RuntimeError(
            "El login mostró la interfaz, pero Playwright no recibió cookies "
            "de sesión para gtd-zcn.etadirect.com."
        )

    print(
        f"[LOGIN] Sesión iniciada correctamente "
        f"({len(session_cookies)} cookies disponibles)."
    )


def _filename_from_content_disposition(
    content_disposition: str,
    fallback_name: str,
) -> str:
    """
    Obtiene un nombre seguro desde Content-Disposition.

    TOA puede informar un nombre .csv aunque el script lo transforme a XLSX.
    Por eso el resultado siempre termina en .xlsx.
    """
    candidate = ""

    if content_disposition:
        match = re.search(
            r"filename\*\s*=\s*(?:UTF-8''|utf-8'')([^;]+)",
            content_disposition,
            flags=re.IGNORECASE,
        )
        if match:
            from urllib.parse import unquote
            candidate = unquote(match.group(1).strip().strip('"'))

        if not candidate:
            match = re.search(
                r'filename\s*=\s*"([^"]+)"|filename\s*=\s*([^;]+)',
                content_disposition,
                flags=re.IGNORECASE,
            )
            if match:
                candidate = (match.group(1) or match.group(2) or "").strip()

    candidate = os.path.basename(candidate.strip().strip('"'))
    candidate = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", candidate)

    if candidate:
        stem, _ = os.path.splitext(candidate)
        candidate = f"{stem}.xlsx"
    else:
        candidate = fallback_name

    if not candidate.lower().endswith(".xlsx"):
        candidate = f"{os.path.splitext(candidate)[0]}.xlsx"

    return candidate


def _validate_xlsx_bytes(content: bytes) -> tuple[bool, str]:
    """Comprueba que el contenido sea realmente un libro XLSX."""
    if not content:
        return False, "respuesta vacía"

    if not content.startswith(b"PK"):
        return False, "la respuesta no comienza con la firma ZIP de un XLSX"

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = set(archive.namelist())
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            missing = required - names
            if missing:
                return False, (
                    "ZIP recibido, pero no corresponde a un XLSX; "
                    f"faltan: {', '.join(sorted(missing))}"
                )

            bad_member = archive.testzip()
            if bad_member:
                return False, f"archivo ZIP/XLSX dañado en: {bad_member}"
    except zipfile.BadZipFile:
        return False, "el contenido tiene firma ZIP, pero está corrupto"
    except Exception as exc:
        return False, f"no fue posible validar el XLSX: {exc!r}"

    return True, ""


def _response_text_preview(content: bytes, limit: int = 700) -> str:
    """Genera una vista breve de respuestas HTML, JSON o texto."""
    if not content:
        return ""

    decoded = content[:20_000].decode("utf-8", errors="replace")
    decoded = re.sub(r"(?is)<script.*?</script>", " ", decoded)
    decoded = re.sub(r"(?is)<style.*?</style>", " ", decoded)
    decoded = re.sub(r"(?s)<[^>]+>", " ", decoded)
    decoded = re.sub(r"\s+", " ", decoded).strip()
    return decoded[:limit]


def _decode_toa_csv(content: bytes) -> tuple[str, str]:
    """
    Decodifica el CSV de TOA.

    Oracle suele responder UTF-8 con BOM, pero se consideran codificaciones
    alternativas para evitar que caracteres acentuados dañen la lectura.
    """
    encodings = []

    if content.startswith(b"\xef\xbb\xbf"):
        encodings.append("utf-8-sig")
    elif content.startswith((b"\xff\xfe", b"\xfe\xff")):
        encodings.extend(["utf-16", "utf-16-le", "utf-16-be"])

    encodings.extend(["utf-8-sig", "utf-8", "cp1252", "latin-1"])

    tried = set()
    last_error = None

    for encoding in encodings:
        if encoding in tried:
            continue
        tried.add(encoding)

        try:
            decoded = content.decode(encoding)
            return decoded, encoding
        except UnicodeDecodeError as exc:
            last_error = exc

    raise ValueError(f"No se pudo decodificar el CSV de TOA: {last_error!r}")


def _normalize_toa_headers(headers: list[str]) -> list[str]:
    """
    Convierte los encabezados CSV en los nombres usados por BBDD_HEADERS.

    La exportación de TOA recibida por HTTP viene en español, mientras que el
    consolidado utiliza varios nombres en inglés. También existen dos columnas
    llamadas 'Nombre'; la primera se asigna a 'Name' y la segunda a 'Nombre'.
    """
    direct_map = {
        "Recurso": "Resource",
        "Fecha": "Date",
        "Tipo de actividad": "Activity Type",
        "Lugar de la Reparación": "Lugar de la Reparación",
        "Reparación en Nodo": "Reparación en Nodo",
        "Reparación en DC": "Reparación en DC",
        "Estado de actividad": "Activity Status",
        "Orden de trabajo": "Work Order",
        "Inicio": "Start",
        "Finalización": "End",
        "Inicio - Fin": "Start - End",
        "Inicio – Fin": "Start - End",
        "Inicio — Fin": "Start - End",
        "Tiempo de viaje": "Traveling Time",
        "Duración": "Duration",
        "Dirección": "Address",
        "Ciudad": "City",
        "Estado": "State",
        "Número de cuenta": "Account Number",
        "Zona de trabajo": "Work Zone",
        "ID de actividad": "Activity ID",
        "Fecha Recepción": "Activity Time of Booking",
        "Fecha de Recepción": "Activity Time of Booking",
    }

    normalized = []
    nombre_count = 0

    for index, raw_header in enumerate(headers):
        header = str(raw_header or "").replace("\ufeff", "").strip()

        if header == "Nombre":
            nombre_count += 1
            header = "Name" if nombre_count == 1 else "Nombre"
        else:
            header = direct_map.get(header, header)

        # Evita encabezados completamente vacíos.
        if not header:
            header = f"Columna_{index + 1}"

        normalized.append(header)

    return normalized


def _detect_csv_dialect(csv_text: str):
    """Detecta coma, punto y coma, tabulación o pipe como separador."""
    sample = csv_text[:20_000]

    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        # El log recibido confirma que TOA usa comas.
        return csv.excel


def _convert_csv_bytes_to_xlsx(
    content: bytes,
    destination: str,
) -> tuple[int, str]:
    """
    Convierte en memoria el CSV devuelto por TOA a un archivo XLSX válido.

    Retorna:
        cantidad de filas de datos, codificación detectada.
    """
    csv_text, encoding = _decode_toa_csv(content)

    # TOA puede anteponer líneas vacías. newline="" evita alterar campos
    # que contengan saltos de línea entre comillas.
    stream = io.StringIO(csv_text, newline="")
    dialect = _detect_csv_dialect(csv_text)
    reader = csv.reader(stream, dialect)

    try:
        raw_headers = next(reader)
    except StopIteration:
        raise ValueError("TOA devolvió un CSV vacío, sin encabezados.")

    headers = _normalize_toa_headers(raw_headers)

    if "Date" not in headers:
        raise ValueError(
            "El CSV fue recibido, pero no contiene la columna Fecha/Date. "
            f"Encabezados detectados: {headers[:15]}"
        )

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="TOA")
    ws.append(headers)

    data_rows = 0
    expected_columns = len(headers)

    for raw_row in reader:
        # Ignora líneas completamente vacías.
        if not raw_row or all(str(value or "").strip() == "" for value in raw_row):
            continue

        row = list(raw_row)

        # Mantiene todas las filas con el mismo número de columnas.
        if len(row) < expected_columns:
            row.extend([""] * (expected_columns - len(row)))
        elif len(row) > expected_columns:
            row = row[:expected_columns]

        ws.append(row)
        data_rows += 1

    wb.save(destination)
    wb.close()

    # Validación final con openpyxl, igual que hará el flujo posterior.
    validation_wb = load_workbook(destination, read_only=True, data_only=True)
    try:
        validation_ws = validation_wb.active
        saved_headers = [
            str(cell.value).strip()
            for cell in validation_ws[1]
            if cell.value is not None
        ]
        if "Date" not in saved_headers:
            raise ValueError(
                "La conversión CSV→XLSX terminó, pero el archivo guardado "
                "no conserva la columna Date."
            )
    finally:
        validation_wb.close()

    return data_rows, encoding


def _looks_like_csv(content_type: str, content: bytes) -> bool:
    """Reconoce los tipos MIME que TOA usa para exportar CSV."""
    content_type = (content_type or "").lower()

    csv_mime_markers = (
        "text/csv",
        "application/csv",
        "comma-separated-values",
        "text/x-comma-separated-values",
    )
    if any(marker in content_type for marker in csv_mime_markers):
        return True

    # Respaldo por contenido cuando el servidor informa octet-stream.
    sample = content[:4_000].lstrip(b"\xef\xbb\xbf\x00\r\n\t ")
    return (
        sample.startswith(b'"')
        and b"," in sample
        and (b"\n" in sample or b"\r" in sample)
    )


def trigger_download(
    page: Page,
    url: str,
    download_dir: str,
    fallback_name: str,
    retries: int = 3,
) -> str:
    """
    Descarga una exportación autenticada de TOA.

    Formatos admitidos:
      1. XLSX real: se guarda directamente.
      2. CSV de TOA: se convierte automáticamente a XLSX.

    Nunca guarda HTML, JSON o texto de error con extensión .xlsx.
    """
    ensure_dir(download_dir)

    try:
        user_agent = page.evaluate("() => navigator.userAgent")
    except Exception:
        user_agent = ""

    request_headers = {
        "Accept": (
            "text/csv,application/csv,"
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet,application/octet-stream;q=0.9,*/*;q=0.8"
        ),
        "Referer": page.url,
    }
    if user_agent:
        request_headers["User-Agent"] = user_agent

    last_error = None

    for attempt in range(1, retries + 1):
        response = None

        try:
            response = page.request.get(
                url,
                headers=request_headers,
                timeout=180_000,
                fail_on_status_code=False,
                max_redirects=10,
            )

            status = response.status
            final_url = response.url
            response_headers = {
                str(key).lower(): str(value)
                for key, value in response.headers.items()
            }
            content_type = response_headers.get("content-type", "").lower()
            content_disposition = response_headers.get(
                "content-disposition",
                "",
            )
            content = response.body()

            # Liberar la respuesta antes de cualquier espera/reintento evita
            # llamadas pendientes si el usuario interrumpe con Ctrl+C.
            try:
                response.dispose()
            except Exception:
                pass
            response = None

            if status < 200 or status >= 300:
                preview = _response_text_preview(content)
                raise RuntimeError(
                    f"HTTP {status}; URL final: {final_url}; "
                    f"Content-Type: {content_type or 'sin informar'}; "
                    f"respuesta: {preview or '<sin contenido>'}"
                )

            file_name = _filename_from_content_disposition(
                content_disposition,
                fallback_name,
            )
            destination = _available_download_path(
                download_dir,
                file_name,
                fallback_name,
            )
            file_name = os.path.basename(destination)
            temporary = os.path.join(
                download_dir,
                f".tmp_{uuid4().hex}_{file_name}",
            )

            is_valid_xlsx, xlsx_error = _validate_xlsx_bytes(content)

            if is_valid_xlsx:
                with open(temporary, "wb") as file:
                    file.write(content)
                    file.flush()
                    os.fsync(file.fileno())

                conversion_description = "XLSX nativo"
                data_rows = None

            elif _looks_like_csv(content_type, content):
                data_rows, encoding = _convert_csv_bytes_to_xlsx(
                    content,
                    temporary,
                )
                conversion_description = (
                    f"CSV convertido a XLSX | codificación {encoding} | "
                    f"{data_rows} filas"
                )

            else:
                preview = _response_text_preview(content)
                raise RuntimeError(
                    f"respuesta no válida como XLSX ni CSV ({xlsx_error}); "
                    f"URL final: {final_url}; "
                    f"Content-Type: {content_type or 'sin informar'}; "
                    f"respuesta: {preview or '<contenido binario desconocido>'}"
                )

            # Validación estructural del archivo definitivo temporal.
            with open(temporary, "rb") as file:
                disk_content = file.read()

            valid_on_disk, disk_error = _validate_xlsx_bytes(disk_content)
            if not valid_on_disk:
                try:
                    os.remove(temporary)
                except OSError:
                    pass
                raise RuntimeError(
                    f"el archivo temporal quedó inválido: {disk_error}"
                )

            os.replace(temporary, destination)

            print(
                f"[DESCARGA HTTP] HTTP {status} | "
                f"{len(content):,} bytes recibidos | "
                f"{conversion_description} | {file_name}"
            )
            return destination

        except KeyboardInterrupt:
            print(
                "[DESCARGA HTTP] Interrupción solicitada por el usuario.",
                file=sys.stderr,
            )
            raise

        except Exception as exc:
            last_error = exc
            print(
                f"[DESCARGA HTTP] Intento {attempt}/{retries} falló: {exc}",
                file=sys.stderr,
            )

            if attempt < retries:
                # No usar page.wait_for_timeout aquí: si se presiona Ctrl+C,
                # Playwright puede dejar corutinas pendientes en Windows.
                time.sleep(1.5 * attempt)

        finally:
            if response is not None:
                try:
                    response.dispose()
                except Exception:
                    pass

    raise RuntimeError(
        f"No fue posible descargar y convertir el archivo después de "
        f"{retries} intentos. Último error: {last_error!r}"
    )


def safe_logout(
    page: Page,
    avatar_xpath: str = "",
    signout_xpath: str = "",
) -> bool:
    """Cierra la sesión TOA usando primero los XPath configurados y luego selectores alternativos."""
    last_error = None

    for _ in range(3):
        try:
            avatar_candidates = []
            if avatar_xpath:
                avatar_candidates.append(xpath_locator(page, avatar_xpath).first)
            avatar_candidates.append(
                page.locator(
                    "xpath=//header//button[.//*[name()='visuals:technician-avatar']]"
                ).first
            )
            avatar_candidates.append(
                page.locator("xpath=//*[name()='visuals:technician-avatar']").first
            )

            avatar_clicked = False
            for avatar in avatar_candidates:
                try:
                    if avatar.count() == 0:
                        continue
                    avatar.wait_for(state="visible", timeout=8_000)
                    avatar.hover()
                    avatar.click(force=True)
                    avatar_clicked = True
                    break
                except Exception as exc:
                    last_error = exc

            if not avatar_clicked:
                raise PlaywrightTimeoutError("No se encontró el avatar de TOA.")

            page.wait_for_timeout(1_000)

            signout_candidates = []
            if signout_xpath:
                signout_candidates.append(xpath_locator(page, signout_xpath).first)
            signout_candidates.extend(
                [
                    page.locator("span.item-caption.item-caption--logout").first,
                    page.locator(
                        "xpath=//span[contains(@class,'item-caption--logout') and "
                        "(normalize-space()='Sign out' "
                        "or normalize-space()='Cerrar sesión')]"
                    ).first,
                    page.locator(
                        "xpath=//a[.//span[normalize-space()='Sign out' "
                        "or normalize-space()='Cerrar sesión']]"
                    ).first,
                ]
            )

            logout_clicked = False
            for logout in signout_candidates:
                try:
                    if logout.count() == 0:
                        continue
                    logout.wait_for(state="visible", timeout=8_000)
                    logout.scroll_into_view_if_needed()
                    logout.click(force=True)
                    logout_clicked = True
                    break
                except Exception as exc:
                    last_error = exc

            if not logout_clicked:
                raise PlaywrightTimeoutError("No se encontró la opción para cerrar sesión.")

            page.wait_for_timeout(1_000)
            return True
        except Exception as exc:
            last_error = exc
            page.wait_for_timeout(1_000)

    raise PlaywrightTimeoutError(
        repr(last_error) if last_error else "No fue posible cerrar sesión"
    )

def enviar_a_power_automate():
    """
    Lee el archivo consolidado final, lo convierte a Base64 y lo envía
    al flujo de Power Automate para su carga en SharePoint.
    """
    ruta_local     = os.getenv("RUTA_BBDD_TOA")
    url_webhook    = os.getenv("PA_WEBHOOK_URL")
    nombre_destino = os.getenv("PA_DESTINATION_NAME")

    if not all([ruta_local, url_webhook, nombre_destino]):
        print("[WEBHOOK] Error: Faltan variables de configuración en .env")
        return False

    if not os.path.exists(ruta_local):
        print(f"[WEBHOOK] Error: No se encontró el archivo para enviar en {ruta_local}")
        return False

    try:
        print(f"[WEBHOOK] Iniciando envío de {nombre_destino}...")
        with open(ruta_local, "rb") as file:
            encoded_string = base64.b64encode(file.read()).decode("utf-8")

        payload = {
            "nombre_archivo": nombre_destino,
            "contenido_base64": encoded_string,
        }

        response = requests.post(url_webhook, json=payload, timeout=120)
        response.raise_for_status()

        print("✅ [WEBHOOK] Archivo enviado exitosamente a Power Automate.")
        return True
    except Exception as e:
        print(f"❌ [WEBHOOK] Error crítico en el envío: {e}")
        return False


# =========================
# PostgreSQL snapshot helpers
# =========================
def env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "si", "sí", "on"}


def safe_path_component(value: str) -> str:
    value = str(value or "sin_nombre").strip()
    value = re.sub(r'[<>:"/\\|?*]+', "_", value)
    value = re.sub(r"\s+", "_", value)
    return value[:100] or "sin_nombre"


def normalize_pg_identifier(value: str) -> str:
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "_", value).strip("_")
    if not value:
        value = "columna"
    if value[0].isdigit():
        value = f"col_{value}"
    return value[:55]


def postgres_columns(headers: list[str]) -> list[str]:
    used: dict[str, int] = {}
    result = []
    for header in headers:
        base = normalize_pg_identifier(header)
        number = used.get(base, 0) + 1
        used[base] = number
        result.append(base if number == 1 else f"{base}_{number}")
    return result


def db_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def pg_connection():
    dsn = os.getenv("PG_DSN")
    if dsn:
        return psycopg.connect(dsn, autocommit=True)

    required = {
        "PG_HOST": os.getenv("PG_HOST"),
        "PG_PORT": os.getenv("PG_PORT", "5432"),
        "PG_DATABASE": os.getenv("PG_DATABASE"),
        "PG_USER": os.getenv("PG_USER"),
        "PG_PASSWORD": os.getenv("PG_PASSWORD"),
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise RuntimeError(f"Faltan variables PostgreSQL: {', '.join(missing)}")

    return psycopg.connect(
        host=required["PG_HOST"],
        port=int(required["PG_PORT"]),
        dbname=required["PG_DATABASE"],
        user=required["PG_USER"],
        password=required["PG_PASSWORD"],
        connect_timeout=int(os.getenv("PG_CONNECT_TIMEOUT", "15")),
        autocommit=True,
    )


def ensure_postgres_objects(conn, headers: list[str]):
    schema = os.getenv("PG_SCHEMA", "public")
    actual = os.getenv("PG_TABLE_ACTUAL", "toa_actual")
    staging = os.getenv("PG_TABLE_STAGING", "toa_staging")
    runs = os.getenv("PG_TABLE_RUNS", "toa_etl_ejecuciones")
    columns = postgres_columns(headers)

    with conn.cursor() as cur:
        cur.execute(sql.SQL("CREATE SCHEMA IF NOT EXISTS {}").format(sql.Identifier(schema)))

        for table_name, unlogged in ((actual, False), (staging, True)):
            create_kw = sql.SQL("CREATE UNLOGGED TABLE") if unlogged else sql.SQL("CREATE TABLE")
            cur.execute(
                sql.SQL("{} IF NOT EXISTS {}.{} (execution_id UUID NOT NULL, loaded_at TIMESTAMPTZ NOT NULL)").format(
                    create_kw,
                    sql.Identifier(schema),
                    sql.Identifier(table_name),
                )
            )
            for column in columns:
                cur.execute(
                    sql.SQL("ALTER TABLE {}.{} ADD COLUMN IF NOT EXISTS {} TEXT").format(
                        sql.Identifier(schema),
                        sql.Identifier(table_name),
                        sql.Identifier(column),
                    )
                )

        cur.execute(
            sql.SQL(
                """
                CREATE TABLE IF NOT EXISTS {}.{} (
                    execution_id UUID PRIMARY KEY,
                    inicio TIMESTAMPTZ NOT NULL,
                    termino TIMESTAMPTZ,
                    archivos_esperados INTEGER NOT NULL,
                    archivos_descargados INTEGER NOT NULL DEFAULT 0,
                    archivos_vacios INTEGER NOT NULL DEFAULT 0,
                    archivos_con_datos INTEGER NOT NULL DEFAULT 0,
                    filas_cargadas BIGINT NOT NULL DEFAULT 0,
                    estado TEXT NOT NULL,
                    observacion TEXT
                )
                """
            ).format(sql.Identifier(schema), sql.Identifier(runs))
        )

    return schema, actual, staging, runs, columns


def register_run(
    conn,
    schema: str,
    runs_table: str,
    execution_id,
    inicio,
    expected_files: int,
    downloaded_files: int,
    empty_files: int,
    data_files: int,
    rows_count: int,
    status: str,
    observation: str = "",
    finished: bool = False,
):
    with conn.cursor() as cur:
        cur.execute(
            sql.SQL(
                """
                INSERT INTO {}.{} (
                    execution_id, inicio, termino, archivos_esperados,
                    archivos_descargados, archivos_vacios, archivos_con_datos,
                    filas_cargadas, estado, observacion
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (execution_id) DO UPDATE SET
                    termino = EXCLUDED.termino,
                    archivos_esperados = EXCLUDED.archivos_esperados,
                    archivos_descargados = EXCLUDED.archivos_descargados,
                    archivos_vacios = EXCLUDED.archivos_vacios,
                    archivos_con_datos = EXCLUDED.archivos_con_datos,
                    filas_cargadas = EXCLUDED.filas_cargadas,
                    estado = EXCLUDED.estado,
                    observacion = EXCLUDED.observacion
                """
            ).format(sql.Identifier(schema), sql.Identifier(runs_table)),
            (
                execution_id,
                inicio,
                now_cl() if finished else None,
                expected_files,
                downloaded_files,
                empty_files,
                data_files,
                rows_count,
                status,
                observation[:10000],
            ),
        )


def publish_postgres_snapshot(
    rows: list,
    headers: list[str],
    execution_id,
    inicio,
    expected_files: int,
    downloaded_files: int,
    empty_files: int,
    data_files: int,
):
    lock_key = int(os.getenv("PG_ADVISORY_LOCK_KEY", "987654321"))
    conn = pg_connection()
    lock_acquired = False

    try:
        schema, actual, staging, runs, columns = ensure_postgres_objects(conn, headers)
        register_run(
            conn, schema, runs, execution_id, inicio,
            expected_files, downloaded_files, empty_files, data_files,
            len(rows), "CARGANDO_STAGING"
        )

        with conn.cursor() as cur:
            cur.execute("SELECT pg_try_advisory_lock(%s)", (lock_key,))
            lock_acquired = bool(cur.fetchone()[0])

        if not lock_acquired:
            register_run(
                conn, schema, runs, execution_id, inicio,
                expected_files, downloaded_files, empty_files, data_files,
                len(rows), "OMITIDA_POR_BLOQUEO",
                "Existe otra ejecución publicando la fotografía de TOA.",
                finished=True,
            )
            raise RuntimeError("Existe otra ejecución TOA publicando datos en PostgreSQL.")

        loaded_at = now_cl()
        copy_columns = columns + ["execution_id", "loaded_at"]
        qualified_staging = sql.SQL("{}.{}").format(sql.Identifier(schema), sql.Identifier(staging))
        qualified_actual = sql.SQL("{}.{}").format(sql.Identifier(schema), sql.Identifier(actual))
        quoted_columns = sql.SQL(", ").join(map(sql.Identifier, copy_columns))

        with conn.transaction():
            with conn.cursor() as cur:
                cur.execute(sql.SQL("TRUNCATE TABLE {}").format(qualified_staging))

                copy_statement = sql.SQL("COPY {} ({}) FROM STDIN").format(
                    qualified_staging,
                    quoted_columns,
                )
                with cur.copy(copy_statement) as copy:
                    for row in rows:
                        values = [db_value(v) for v in row]
                        values.extend([execution_id, loaded_at])
                        copy.write_row(values)

                cur.execute(sql.SQL("SELECT COUNT(*) FROM {}").format(qualified_staging))
                staged_count = int(cur.fetchone()[0])
                if staged_count != len(rows):
                    raise RuntimeError(
                        f"Validación staging fallida: esperadas {len(rows)} filas, cargadas {staged_count}."
                    )

                cur.execute(sql.SQL("LOCK TABLE {} IN ACCESS EXCLUSIVE MODE").format(qualified_actual))
                cur.execute(sql.SQL("TRUNCATE TABLE {}").format(qualified_actual))
                cur.execute(
                    sql.SQL("INSERT INTO {} ({}) SELECT {} FROM {}").format(
                        qualified_actual,
                        quoted_columns,
                        quoted_columns,
                        qualified_staging,
                    )
                )

        register_run(
            conn, schema, runs, execution_id, inicio,
            expected_files, downloaded_files, empty_files, data_files,
            len(rows), "EXITOSA", finished=True
        )
        print(f"[POSTGRESQL] Fotografía publicada: {len(rows)} filas.")
        return len(rows)

    except Exception as exc:
        try:
            if 'schema' in locals():
                register_run(
                    conn, schema, runs, execution_id, inicio,
                    expected_files, downloaded_files, empty_files, data_files,
                    len(rows), "ERROR_POSTGRESQL", repr(exc), finished=True
                )
        except Exception:
            pass
        raise
    finally:
        if lock_acquired:
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT pg_advisory_unlock(%s)", (lock_key,))
            except Exception:
                pass
        conn.close()


def consume_downloaded_xlsx(
    xlsx_path: str,
    subgroup: str,
    zona_subgrupo: str,
    headers: list[str],
) -> pd.DataFrame:
    """
    Lee el Excel, valida sus encabezados exactos y filtra Activity Type.

    El archivo no se elimina aquí: la limpieza se realiza únicamente cuando
    la ejecución completa termina correctamente.
    """
    try:
        source_df = pd.read_excel(xlsx_path, engine="openpyxl")
    except Exception as exc:
        raise RuntimeError(
            f"No se pudo leer el archivo '{xlsx_path}': {exc!r}"
        ) from exc

    required_columns = ["Date", "Activity Type"]
    missing_columns = [
        column for column in required_columns if column not in source_df.columns
    ]
    if missing_columns:
        raise ValueError(
            f"Archivo sin columna(s) obligatoria(s) con nombre exacto "
            f"{missing_columns}: {xlsx_path}. "
            f"Encabezados detectados: {list(source_df.columns)}"
        )

    filtered_df = source_df[
        source_df["Activity Type"].isin(ACTIVITY_TYPE_ALLOWED_VALUES)
    ].copy()
    filtered_df["Subgrupo TOA"] = subgroup
    filtered_df["Zona Subgrupo TOA"] = zona_subgrupo
    filtered_df = filtered_df.reindex(columns=headers)

    print(
        f"[FILTRO] {os.path.basename(xlsx_path)} | "
        f"registros origen: {len(source_df)} | "
        f"registros permitidos: {len(filtered_df)}"
    )
    return filtered_df


def dataframe_to_rows(dataframe: pd.DataFrame) -> list[list]:
    """Convierte el consolidado a valores compatibles con Excel y PostgreSQL."""
    rows = []
    for source_row in dataframe.itertuples(index=False, name=None):
        row = []
        for value in source_row:
            if pd.isna(value):
                row.append(None)
            elif isinstance(value, pd.Timestamp):
                row.append(value.to_pydatetime())
            elif isinstance(value, pd.Timedelta):
                row.append(value.to_pytimedelta())
            elif hasattr(value, "item"):
                row.append(value.item())
            else:
                row.append(value)
        rows.append(row)
    return rows


def write_snapshot_excel(xlsx_path: str, headers: list[str], rows: list):
    ensure_dir(os.path.dirname(xlsx_path) or ".")
    wb = Workbook()
    ws = wb.active
    ws.title = "BBDD_TOA"

    for column_number, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column_number, value=header)

    for row_number, row in enumerate(rows, start=2):
        for column_number, value in enumerate(row, start=1):
            ws.cell(row=row_number, column=column_number, value=value)

    end_col = get_column_letter(len(headers))
    end_row = max(1, len(rows) + 1)
    table = Table(displayName="BBDD_TOA", ref=f"A1:{end_col}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    temp_path = f"{xlsx_path}.tmp.xlsx"
    wb.save(temp_path)
    wb.close()
    os.replace(temp_path, xlsx_path)


def save_manifest(run_dir: str, execution_id, inicio, expected_files: int, manifest: list, errors: list):
    payload = {
        "execution_id": str(execution_id),
        "inicio": inicio.isoformat(),
        "archivos_esperados": expected_files,
        "archivos_descargados": len(manifest),
        "errores": errors,
        "archivos": manifest,
    }
    manifest_path = os.path.join(run_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    return manifest_path


# =========================
# MAIN seguro: descarga aislada + fotografía PostgreSQL
# =========================
def run_process(cleanup_state: dict):
    load_dotenv()

    USER = os.getenv("TOA_USERNAME")
    PASS = os.getenv("TOA_PASSWORD")

    # XPath del proyecto original. Se pueden sobrescribir desde .env.
    X_USER = (
        os.getenv("TOA_XPATH_USER")
        or "/html/body/div/div/div/div/form/div[2]/div[1]/div[1]/div/input"
    )
    X_PASS = (
        os.getenv("TOA_XPATH_PASS")
        or "/html/body/div/div/div/div/form/div[2]/div[2]/div[1]/div/input"
    )
    X_BTN = (
        os.getenv("TOA_XPATH_BTN")
        or "/html/body/div/div/div/div/form/div[2]/div[5]/button"
    )
    X_BTN_ALT = (
        os.getenv("TOA_XPATH_BTN_ALT")
        or "/html/body/div/div/div/div/form/div[2]/div[6]/button"
    )
    X_READY = (
        os.getenv("TOA_XPATH_READY")
        or "/html/body/div[14]/div[1]/main/div/div[2]/div[3]/div[1]/div[3]/div[1]/div[2]/div[2]/div/div[2]/div[3]/div[2]/div[2]/div[5]/div[1]/button[3]/span"
    )
    X_AVATAR = (
        os.getenv("TOA_XPATH_AVATAR")
        or "/html/body/div[14]/div[1]/app:global-header/header/div[5]/button/visuals:technician-avatar"
    )
    X_SIGNOUT = (
        os.getenv("TOA_XPATH_SIGNOUT")
        or "/html/body/div[26]/div[2]/app:global-header/div/ul/li[4]/a"
    )

    RUTA_EXC = os.getenv("RUTA_EXC_MONITOREO")
    RUTA_BBDD_TOA = os.getenv("RUTA_BBDD_TOA")
    RUTA_INPUTS = os.getenv("RUTA_INPUTS_TOA")
    RUNS_DIR = os.getenv("TOA_RUNS_DIR") or os.path.join(os.path.dirname(RUTA_BBDD_TOA or "."), "toa_runs")
    DOWNLOAD_DIR = os.getenv("DOWNLOAD_DIR") or os.path.join(RUNS_DIR, "downloads")
    SHEET_EXC = "Ejecuciones"
    TABLE_EXC = "Ejecuciones"

    required = {
        "TOA_USERNAME": USER,
        "TOA_PASSWORD": PASS,
        "RUTA_EXC_MONITOREO": RUTA_EXC,
        "RUTA_BBDD_TOA": RUTA_BBDD_TOA,
        "RUTA_INPUTS_TOA": RUTA_INPUTS,
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise RuntimeError(f"Faltan variables en .env: {', '.join(missing)}")

    ensure_dir(DOWNLOAD_DIR)
    ensure_dir(RUNS_DIR)
    remove_old_download_files(
        DOWNLOAD_DIR,
        protected_paths=[RUTA_EXC, RUTA_BBDD_TOA, RUTA_INPUTS],
    )

    inputs_records = load_inputs_toa(RUTA_INPUTS)
    if not inputs_records:
        raise RuntimeError("No se encontraron registros ACTIVOS en InputsTOA.xlsx.")

    execution_id = uuid4()
    inicio = now_cl()
    run_dir = os.path.join(RUNS_DIR, f"{inicio.strftime('%Y%m%d_%H%M%S')}_{execution_id}")
    ensure_dir(run_dir)

    days_back = int(os.getenv("TOA_DAYS_BACK", "3"))
    days_forward = int(os.getenv("TOA_DAYS_FORWARD", "14"))
    base_date = now_cl().date()
    date_range = [
        base_date + timedelta(days=offset)
        for offset in range(-days_back, days_forward + 1)
    ]
    expected_files = len(date_range) * len(inputs_records)

    counts: dict[str, int] = {record["Subgrupo TOA"]: 0 for record in inputs_records}
    manifest: list[dict] = []
    download_errors: list[str] = []
    non_blocking_errors: list[str] = []

    dataframes_filtrados: list[pd.DataFrame] = []
    empty_files = 0
    data_files = 0
    rows_by_subgroup: dict[str, int] = {}

    headless = env_bool("TOA_HEADLESS", False)
    browser_channel = (os.getenv("TOA_BROWSER_CHANNEL", "msedge") or "").strip()
    slow_mo = int(os.getenv("TOA_SLOW_MO_MS", "0"))

    playwright = None
    browser = None
    context: BrowserContext | None = None
    page: Page | None = None

    try:
        playwright = sync_playwright().start()

        launch_kwargs = {
            "headless": headless,
            "slow_mo": slow_mo,
        }
        # Vacío, "chromium" o "bundled" usa Chromium administrado por Playwright.
        if browser_channel.lower() not in {"", "chromium", "bundled"}:
            launch_kwargs["channel"] = browser_channel
        if not headless:
            launch_kwargs["args"] = ["--start-maximized"]

        browser = playwright.chromium.launch(**launch_kwargs)
        context = browser.new_context(
            accept_downloads=True,
            viewport=None if not headless else {"width": 1920, "height": 1080},
            locale="es-CL",
            timezone_id="America/Santiago",
        )
        context.set_default_timeout(60_000)
        context.set_default_navigation_timeout(120_000)
        page = context.new_page()

        login_toa(
            page=page,
            username=USER,
            password=PASS,
            user_xpath=X_USER,
            password_xpath=X_PASS,
            button_xpath=X_BTN,
            button_xpath_alt=X_BTN_ALT,
            ready_xpath=X_READY,
        )

        for current_date in date_range:
            date_str = current_date.strftime("%Y-%m-%d")
            print(f"[DESCARGA] Fecha {date_str}")

            for index, record in enumerate(inputs_records, start=1):
                subgrupo = record["Subgrupo TOA"]
                zona_subgrupo = record["Zona Subgrupo TOA"]
                url = build_toa_url(record, date_str)

                try:
                    fallback_name = (
                        f"{date_str}_{safe_path_component(subgrupo)}_"
                        f"{safe_path_component(record.get('downloadId', str(index)))}.xlsx"
                    )
                    downloaded_path = trigger_download(
                        page,
                        url,
                        DOWNLOAD_DIR,
                        fallback_name=fallback_name,
                    )
                    cleanup_state["downloaded_files"].append(downloaded_path)
                    file_name = os.path.basename(downloaded_path)
                    print(
                        f"[DESCARGA] Archivo descargado correctamente: "
                        f"{downloaded_path}"
                    )

                    filtered_df = consume_downloaded_xlsx(
                        downloaded_path,
                        subgrupo,
                        zona_subgrupo,
                        BBDD_HEADERS,
                    )

                    dataframes_filtrados.append(filtered_df)
                    filtered_rows_count = len(filtered_df)
                    if filtered_rows_count:
                        data_files += 1
                        rows_by_subgroup[subgrupo] = (
                            rows_by_subgroup.get(subgrupo, 0)
                            + filtered_rows_count
                        )
                    else:
                        empty_files += 1

                    manifest.append(
                        {
                            "file_name": file_name,
                            "date": date_str,
                            "subgrupo": subgrupo,
                            "zona_subgrupo": zona_subgrupo,
                            "download_id": record.get("downloadId", ""),
                            "provider_id": record.get("providerId", ""),
                            "rows": filtered_rows_count,
                            "empty": filtered_rows_count == 0,
                            "retained_until_process_end": True,
                        }
                    )
                    counts[subgrupo] = counts.get(subgrupo, 0) + 1
                    print(
                        f"[DESCARGA] {subgrupo} OK → {file_name} "
                        f"| registros filtrados: {filtered_rows_count}"
                    )
                    page.wait_for_timeout(400)
                except Exception as exc:
                    error = f"[{date_str}][{subgrupo}] {exc!r}"
                    download_errors.append(error)
                    print(f"[DESCARGA] ERROR {error}", file=sys.stderr)

        try:
            safe_logout(
                page,
                avatar_xpath=X_AVATAR,
                signout_xpath=X_SIGNOUT,
            )
        except Exception as exc:
            non_blocking_errors.append(f"[LOGOUT] {exc!r}")

    except Exception as exc:
        message = repr(exc)
        if "Executable doesn't exist" in message or "browserType.launch" in message:
            message += (
                " | Instala el navegador con: py -m playwright install msedge "
                "o deja TOA_BROWSER_CHANNEL vacío y ejecuta: "
                "py -m playwright install chromium"
            )
        download_errors.append(f"[FATAL] {message}")
    finally:
        if context is not None:
            try:
                context.close()
            except Exception:
                pass
        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass
        if playwright is not None:
            try:
                playwright.stop()
            except Exception:
                pass

    manifest_path = save_manifest(
        run_dir,
        execution_id,
        inicio,
        expected_files,
        manifest,
        download_errors + non_blocking_errors,
    )
    print(f"[MANIFEST] {manifest_path}")

    # El monitoreo Excel se mantiene por compatibilidad.
    fin_descarga = now_cl()
    result_download = "Exitosa" if not download_errors and len(manifest) == expected_files else "Con Errores"
    observations = "\n".join(download_errors + non_blocking_errors)
    counts_summary = " | ".join(f"{sg}: {count}" for sg, count in counts.items())

    try:
        row = [
            human_dt(inicio),
            human_dt(fin_descarga),
            *[counts.get(record["Subgrupo TOA"], 0) for record in inputs_records],
            result_download,
            observations,
        ]
        append_to_table_atomic(RUTA_EXC, SHEET_EXC, TABLE_EXC, row)
    except Exception as exc:
        non_blocking_errors.append(f"[MONITOREO EXCEL] {exc!r}")
        print(f"[MONITOREO] Error: {exc!r}", file=sys.stderr)

    print(f"[RESUMEN] Esperados: {expected_files} | Descargados: {len(manifest)}")
    print(f"[RESUMEN] {counts_summary}")

    # Una foto incompleta nunca reemplaza la base vigente.
    if download_errors or len(manifest) != expected_files:
        error_message = (
            f"Fotografía incompleta: {len(manifest)}/{expected_files} archivos. "
            "PostgreSQL no fue modificado."
        )
        try:
            with pg_connection() as conn:
                schema, _, _, runs, _ = ensure_postgres_objects(conn, BBDD_HEADERS)
                register_run(
                    conn, schema, runs, execution_id, inicio,
                    expected_files, len(manifest), 0, 0, 0,
                    "ERROR_DESCARGA",
                    error_message + "\n" + "\n".join(download_errors),
                    finished=True,
                )
        except Exception as exc:
            print(f"[POSTGRESQL] No se pudo registrar el error: {exc!r}", file=sys.stderr)

        try:
            update_last_row_fields_atomic(
                RUTA_EXC, SHEET_EXC, TABLE_EXC,
                {
                    "Resultado_Consolidado": "No publicado",
                    "Obs_Consolidado": error_message,
                },
            )
        except Exception:
            pass
        raise RuntimeError(error_message)

    if dataframes_filtrados:
        df_consolidado = pd.concat(dataframes_filtrados, ignore_index=True)
        df_consolidado = df_consolidado.reindex(columns=BBDD_HEADERS)
    else:
        df_consolidado = pd.DataFrame(columns=BBDD_HEADERS)

    rows = dataframe_to_rows(df_consolidado)
    print(f"[CONSOLIDADO] Total de registros: {len(df_consolidado)}")

    print(
        f"[CONSOLIDADO] Archivos con datos: {data_files} | "
        f"Archivos vacíos: {empty_files} | Filas: {len(rows)}"
    )
    for subgroup, row_count in rows_by_subgroup.items():
        print(f"  - {subgroup}: {row_count} filas")

    # El Excel es un artefacto de salida; la publicación PostgreSQL usa las filas en memoria.
    write_snapshot_excel(RUTA_BBDD_TOA, BBDD_HEADERS, rows)

    postgres_ok = False
    postgres_error = ""
    try:
        publish_postgres_snapshot(
            rows,
            BBDD_HEADERS,
            execution_id,
            inicio,
            expected_files,
            len(manifest),
            empty_files,
            data_files,
        )
        postgres_ok = True
    except Exception as exc:
        postgres_error = repr(exc)
        print(f"[POSTGRESQL] ERROR: {postgres_error}", file=sys.stderr)

    webhook_ok = False
    if postgres_ok and env_bool("PA_ENABLED", True):
        webhook_ok = enviar_a_power_automate()
    elif postgres_ok:
        webhook_ok = True

    try:
        update_last_row_fields_atomic(
            RUTA_EXC,
            SHEET_EXC,
            TABLE_EXC,
            {
                "Resultado_Consolidado": "Exitoso" if postgres_ok else "Error PostgreSQL",
                "Obs_Consolidado": "" if postgres_ok else postgres_error,
            },
        )
    except Exception as exc:
        print(f"[MONITOREO] No se pudo actualizar consolidado: {exc!r}", file=sys.stderr)

    if not postgres_ok:
        raise RuntimeError(f"No fue posible publicar PostgreSQL: {postgres_error}")

    # Los archivos descargados se eliminan al finalizar satisfactoriamente.
    # La carpeta de ejecución conserva manifest.json para auditoría.
    if webhook_ok and env_bool("CLEAN_RUN_MANIFEST", False):
        try:
            shutil.rmtree(run_dir)
            print("[LIMPIEZA] Carpeta de manifiesto eliminada.")
        except Exception as exc:
            print(f"[LIMPIEZA] No se pudo eliminar {run_dir}: {exc!r}", file=sys.stderr)

    cleanup_state["proceso_completado"] = True


def main():
    cleanup_state = {
        "proceso_completado": False,
        "downloaded_files": [],
    }

    try:
        run_process(cleanup_state)
    except Exception as error:
        print(
            f"[ERROR PROCESO] Error durante el procesamiento: {error!r}",
            file=sys.stderr,
        )
        raise
    finally:
        downloaded_files = cleanup_state["downloaded_files"]
        if cleanup_state["proceso_completado"]:
            cleanup_errors = _remove_files(
                downloaded_files,
                "[LIMPIEZA FINAL]",
            )
            if cleanup_errors:
                print(
                    "[LIMPIEZA FINAL] El proceso terminó correctamente, pero "
                    "algunos archivos no pudieron eliminarse.",
                    file=sys.stderr,
                )
        elif downloaded_files:
            print(
                "[LIMPIEZA FINAL] El proceso no terminó correctamente; "
                "se conservan los archivos descargados para revisión:"
            )
            for path in downloaded_files:
                if os.path.exists(path):
                    print(f"[LIMPIEZA FINAL] Conservado: {path}")

    if cleanup_state["proceso_completado"]:
        print("---- EJECUCIÓN COMPLETA Y PUBLICADA ----")


if __name__ == "__main__":
    main()
